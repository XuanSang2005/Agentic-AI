"""Load POI_Dataset + Public_Evaluation thành POI documents — nguồn chọn qua env.

Nguồn sự thật duy nhất về data — mọi lớp trên (BM25, dense, filter, rerank, eval)
đọc qua module này, không tự mở xlsx / query DB.

POI đọc được từ 2 nguồn TƯƠNG ĐƯƠNG (equivalence-gated, tests/test_postgres_equivalence.py):
- DATA_SOURCE=xlsx (mặc định): data/*.xlsx qua openpyxl — offline, không cần gì thêm.
- DATA_SOURCE=postgres (opt-in): bảng pois (migrations/001_init.sql, seed bằng
  scripts/seed_postgres.py). ORDER BY row_order = vị trí dòng xlsx gốc — thứ tự
  corpus quyết định hash embedding cache, KHÔNG được đổi. document/norm_document/
  is_synthetic luôn derive bằng CÙNG code path (_finalize) bất kể nguồn.
Public_Evaluation luôn đọc từ xlsx — eval harness là tooling offline, ngoài scope DB.

- GIỮ NGUYÊN dấu tiếng Việt trong document lưu trữ; chỉ bỏ dấu ở tầng matcher
  (normalize_vi là util chuẩn hóa dùng chung cho MỌI bước so khớp).
- Text field để BM25/dense dùng chung được dựng sẵn Ở ĐÂY (document / norm_document)
  — tránh mỗi retriever tự ghép một kiểu rồi lệch nhau.
- ĐỪNG xóa dòng G — chỉ đánh dấu is_synthetic để eval chẩn đoán và L3 down-rank.
  TUYỆT ĐỐI không dùng is_synthetic để lọc trong retriever.
"""
from __future__ import annotations

import unicodedata
from dataclasses import dataclass, field
from functools import lru_cache

import openpyxl

from src import config


def normalize_vi(text: str) -> str:
    """Chuẩn hóa tiếng Việt để SO KHỚP: lowercase + bỏ dấu (NFD, bỏ combining marks, đ→d).

    Chỉ dùng khi match — bản gốc có dấu luôn được giữ để hiển thị/embed.
    Nhờ bỏ dấu cả 2 phía (query lẫn document), câu không dấu vẫn match được.
    """
    text = text.lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    # NFD không tách được đ/Đ — thay tay sau khi đã lowercase
    return text.replace("đ", "d")


@dataclass
class POI:
    id: str
    name: str
    brand: str
    category: str
    sub_category: str
    city: str
    district: str
    address: str
    lat: float
    lon: float
    rating: float
    review_count: int
    popularity: float
    price_level: int
    opening_hours: str
    attributes: list[str] = field(default_factory=list)
    tags: list[str] = field(default_factory=list)
    description: str = ""
    is_synthetic: bool = False  # id bắt đầu bằng "G" — CHỈ để chẩn đoán trong eval
    # Trạng thái verify (Phase 4b) — CHỈ để HIỂN THỊ (policy A), không lọc/không
    # vào ranking/document. xlsx không có cột này → default "active" (khớp
    # default cột Postgres — equivalence 2 nguồn giữ nguyên); POI ingest qua
    # /admin mang verified/unverified thật.
    status: str = "active"
    # Text field dựng sẵn cho retrieval — BM25 và dense sau này DÙNG CHUNG:
    document: str = ""       # giữ nguyên dấu — để hiển thị / embed
    norm_document: str = ""  # normalize_vi(document) — để BM25 match


@dataclass
class EvalQuery:
    id: str
    query: str
    category: str      # query_category (Semantic Search, Intent Search, ...)
    difficulty: str    # Easy / Medium / Hard
    expected_ids: list[str] = field(default_factory=list)
    # Giữ RAW string — đừng tokenize theo dấu ("24/7" bị cắt sai); parse ra slot để sau.
    expected_requirements: str = ""
    ranking_signals: list[str] = field(default_factory=list)


def _rows_of(ws) -> list[dict]:
    """Sheet → list[dict] theo header dòng 1, bỏ dòng trống."""
    rows = ws.iter_rows(values_only=True)
    header = [str(h) if h is not None else "" for h in next(rows)]
    out = []
    for row in rows:
        if all(v is None for v in row):
            continue
        out.append(dict(zip(header, row)))
    return out


def _text(value) -> str:
    """Cell → str đã strip; None → chuỗi rỗng (giữ nguyên dấu)."""
    return "" if value is None else str(value).strip()


def _split_list(value, sep: str = ";") -> list[str]:
    """Cell "a;b;c" → ["a", "b", "c"] (strip từng phần, bỏ phần rỗng)."""
    return [p.strip() for p in _text(value).split(sep) if p.strip()]


@lru_cache(maxsize=1)
def _load_workbook_rows() -> dict[str, list[dict]]:
    """Đọc cả 5 sheet một lần, cache lại (xlsx chỉ ~34KB)."""
    wb = openpyxl.load_workbook(config.DATA_XLSX, read_only=True, data_only=True)
    return {name: _rows_of(wb[name]) for name in wb.sheetnames}


def is_synthetic_id(poi_id: str) -> bool:
    """Luật nhận diện dòng G synthetic — NGUỒN DUY NHẤT (loader + ingestion dùng chung)."""
    return poi_id.startswith("G")


def _finalize(poi: POI) -> POI:
    """Derived fields — CODE PATH DUY NHẤT cho mọi nguồn (xlsx/postgres): đổi cách
    ghép document ở đây là đổi hash embedding cache, phải re-encode."""
    poi.is_synthetic = is_synthetic_id(poi.id)
    # Document ghép các field ngữ nghĩa (không address/brand — để location/brand
    # xử lý có chủ đích ở L1/L2, không nhiễu lexical).
    poi.document = " ".join(part for part in [
        poi.name,
        poi.category,
        poi.sub_category,
        poi.district,
        poi.city,
        " ".join(poi.attributes),
        " ".join(poi.tags),
        poi.description,
    ] if part)
    poi.norm_document = normalize_vi(poi.document)
    return poi


def _pois_from_xlsx() -> list[POI]:
    """POI_Dataset (xlsx) → list[POI], giữ nguyên thứ tự dòng, KHÔNG lọc dòng G."""
    return [_finalize(POI(
        id=_text(r["poi_id"]),
        name=_text(r["poi_name"]),
        brand=_text(r["brand"]),
        category=_text(r["category"]),
        sub_category=_text(r["sub_category"]),
        city=_text(r["city"]),
        district=_text(r["district"]),
        address=_text(r["address"]),
        lat=float(r["latitude"] or 0.0),
        lon=float(r["longitude"] or 0.0),
        rating=float(r["rating"] or 0.0),
        review_count=int(r["review_count"] or 0),
        popularity=float(r["popularity_score"] or 0.0),
        price_level=int(r["price_level"] or 0),
        opening_hours=_text(r["opening_hours"]),
        attributes=_split_list(r["attributes"]),
        tags=_split_list(r["tags"]),
        description=_text(r["description"]),
    )) for r in _load_workbook_rows()[config.SHEET_POI]]


# Cột SELECT theo đúng thứ tự field POI; ORDER BY row_order (vị trí dòng xlsx gốc)
# — KHÔNG phải poi_id: xlsx xếp C→R→H→A→S→M→G, sort id sẽ đảo corpus và phá
# hash embedding cache dù data y hệt.
_PG_QUERY = """
    SELECT poi_id, name, brand, category, sub_category, city, district, address,
           lat, lon, rating, review_count, popularity_score, price_level,
           opening_hours, attributes, tags, description, status
    FROM pois
    ORDER BY row_order
"""


@lru_cache(maxsize=1)
def _load_pg_rows() -> tuple[tuple, ...]:
    """Fetch bảng pois đúng 1 lần/process (mirror _load_workbook_rows phía xlsx)."""
    import psycopg  # import tại chỗ: nhánh xlsx không cần driver

    with psycopg.connect(config.database_url()) as conn:
        return tuple(conn.execute(_PG_QUERY).fetchall())


def _pois_from_postgres() -> list[POI]:
    """Bảng pois → list[POI] — ép kiểu Y HỆT nhánh xlsx: NULL xử như ô rỗng
    (_text → ""), số về float/int Python (float8 round-trip đúng bit, không Decimal),
    jsonb array → list[str]."""
    return [_finalize(POI(
        id=_text(r[0]),
        name=_text(r[1]),
        brand=_text(r[2]),
        category=_text(r[3]),
        sub_category=_text(r[4]),
        city=_text(r[5]),
        district=_text(r[6]),
        address=_text(r[7]),
        lat=float(r[8] or 0.0),
        lon=float(r[9] or 0.0),
        rating=float(r[10] or 0.0),
        review_count=int(r[11] or 0),
        popularity=float(r[12] or 0.0),
        price_level=int(r[13] or 0),
        opening_hours=_text(r[14]),
        attributes=[str(a) for a in (r[15] or [])],
        tags=[str(t) for t in (r[16] or [])],
        description=_text(r[17]),
        status=_text(r[18]),
    )) for r in _load_pg_rows()]


def load_pois() -> list[POI]:
    """list[POI] Y HỆT bất kể nguồn — chọn qua env DATA_SOURCE (xlsx | postgres)."""
    if config.DATA_SOURCE == "postgres":
        return _pois_from_postgres()
    return _pois_from_xlsx()


def extract_unique_attributes(pois: list[POI]) -> list[str]:
    """Tập hợp tất cả attribute strings DUY NHẤT trong dataset, sort ổn định cho cache key."""
    attrs: set[str] = set()
    for poi in pois:
        attrs.update(poi.attributes)
    return sorted(attrs)


def extract_unique_categories(pois: list[POI]) -> list[str]:
    """Tập hợp tất cả category strings DUY NHẤT trong dataset, sort ổn định cho cache key."""
    cats: set[str] = set()
    for poi in pois:
        if poi.category:
            cats.add(poi.category)
    return sorted(cats)


def current_data_version() -> int:
    """data_version dùng chung trong Postgres (Phase 5) — KHÔNG cache: poller
    đọc trực tiếp mỗi lần. Chưa có row (DB trước migration 002) → 0."""
    import psycopg

    with psycopg.connect(config.database_url()) as conn:
        row = conn.execute("SELECT value FROM meta WHERE key = 'data_version'").fetchone()
        return int(row[0]) if row else 0


def load_eval() -> list[EvalQuery]:
    """Public_Evaluation → list[EvalQuery]. expected_ids tách ";", signals tách ","."""
    return [
        EvalQuery(
            id=_text(r["query_id"]),
            query=_text(r["input_query"]),
            category=_text(r["query_category"]),
            difficulty=_text(r["difficulty"]),
            expected_ids=_split_list(r["expected_top_poi_ids"]),
            expected_requirements=_text(r["expected_semantic_requirements"]),
            ranking_signals=_split_list(r["ranking_signals_to_use"], sep=","),
        )
        for r in _load_workbook_rows()[config.SHEET_EVAL]
    ]
