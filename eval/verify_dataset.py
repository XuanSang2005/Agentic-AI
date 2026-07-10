"""Xác minh các con số về dataset trước khi tin vào chúng.

Chạy: .venv/bin/python eval/verify_dataset.py
Mọi claim trong CLAUDE.md về dataset phải được kiểm chứng bằng script này.
"""
from collections import Counter
from pathlib import Path

import openpyxl

XLSX = Path(__file__).resolve().parent.parent / "data" / "ai_maps_track2_dataset_participants.xlsx"


def rows_of(ws):
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        out.append(dict(zip(header, row)))
    return out


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    print("Sheets:", wb.sheetnames)

    pois = rows_of(wb["POI_Dataset"])
    queries = rows_of(wb["Public_Evaluation"])
    signals = rows_of(wb["Ranking_Signals"])
    taxonomy = rows_of(wb["Attribute_Taxonomy"])

    # --- POI_Dataset ---
    print(f"\n[POI_Dataset] tổng dòng: {len(pois)}")
    prefix = Counter(p["poi_id"][0] for p in pois)
    print("  theo prefix:", dict(sorted(prefix.items())))
    real = [p for p in pois if not p["poi_id"].startswith("G")]
    synth = [p for p in pois if p["poi_id"].startswith("G")]
    print(f"  THẬT (C/R/H/A/S/M): {len(real)} | SYNTHETIC (G): {len(synth)}")

    # 82 attribute token thật (chỉ tính POI thật)
    tok_real = Counter()
    tok_all = Counter()
    for p in pois:
        toks = [t.strip() for t in str(p["attributes"] or "").split(";") if t.strip()]
        tok_all.update(toks)
        if not p["poi_id"].startswith("G"):
            tok_real.update(toks)
    print(f"  attribute token duy nhất — POI thật: {len(tok_real)} | toàn bộ: {len(tok_all)}")

    # G mâu thuẫn: "gần biển" ở thành phố không có biển
    no_sea = {"Hà Nội", "Đà Lạt", "TP.HCM"}
    bad_g = [p["poi_id"] for p in synth
             if "gần biển" in str(p["attributes"]) and p["city"] in no_sea]
    print(f"  G có 'gần biển' ở city không có biển: {len(bad_g)} → {bad_g}")

    # --- Attribute_Taxonomy / Ranking_Signals ---
    print(f"\n[Attribute_Taxonomy] số nhãn: {len(taxonomy)}")
    print(f"[Ranking_Signals] số signal: {len(signals)} → {[s['signal'] for s in signals]}")

    # --- Public_Evaluation ---
    print(f"\n[Public_Evaluation] tổng câu: {len(queries)}")
    print("  difficulty:", dict(Counter(q["difficulty"] for q in queries)))
    print("  query_category:", dict(Counter(q["query_category"] for q in queries)))

    n_single = sum(1 for q in queries
                   if len(str(q["expected_top_poi_ids"]).split(";")) == 1)
    print(f"  câu chỉ có 1 đáp án: {n_single}/{len(queries)}")

    n_loc = sum(1 for q in queries if "location" in str(q["ranking_signals_to_use"]))
    print(f"  câu có 'location' trong ranking_signals_to_use: {n_loc}/{len(queries)}")

    # Sanity: không dòng G nào là đáp án
    g_ids = {p["poi_id"] for p in synth}
    expected = set()
    for q in queries:
        expected.update(x.strip() for x in str(q["expected_top_poi_ids"]).split(";"))
    leaked = expected & g_ids
    missing = expected - {p["poi_id"] for p in pois}
    print(f"  G-id xuất hiện trong đáp án: {sorted(leaked) or 'KHÔNG'}")
    print(f"  expected_id không tồn tại trong POI_Dataset: {sorted(missing) or 'KHÔNG'}")


if __name__ == "__main__":
    main()
